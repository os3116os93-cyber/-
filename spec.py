import streamlit as st
import pandas as pd
import os
import base64
from openpyxl import load_workbook

try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "admin1234")
EXCEL_FILE = "customer.xlsx"

st.set_page_config(
    page_title="고객사양서 - 품질기술팀",
    page_icon="📋",
    layout="wide"
)

if "is_admin" not in st.session_state:
    st.session_state.is_admin = False
if "edit_idx" not in st.session_state:
    st.session_state.edit_idx = None
if "show_add_form" not in st.session_state:
    st.session_state.show_add_form = False


@st.cache_data(ttl=300)
def get_image_base64(file_path):
    if not os.path.exists(file_path):
        return None
    try:
        with open(file_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    except Exception as e:
        st.error(f"이미지 로드 오류: {e}")
        return None


@st.cache_data(ttl=300)
def load_data(file_name, skip=0):
    file_path = os.path.join(BASE_DIR, file_name)
    if not os.path.exists(file_path):
        st.error(f"파일을 찾을 수 없습니다: {file_name}")
        return None
    try:
        df = pd.read_excel(file_path, engine="openpyxl", skiprows=skip)
        df = df[~df.iloc[:, 0].astype(str).str.contains("※", na=False)]
        return df
    except Exception as e:
        st.error(f"{file_name} 로드 오류: {e}")
        return None


@st.cache_data(ttl=300)
def load_standard_with_merge(file_name, skip=5):
    """openpyxl로 병합셀 정보까지 읽어서 HTML 테이블 생성"""
    file_path = os.path.join(BASE_DIR, file_name)
    if not os.path.exists(file_path):
        st.error(f"파일을 찾을 수 없습니다: {file_name}")
        return None
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # 병합셀 범위 수집
        merge_map = {}   # (row, col) -> (rowspan, colspan)
        skip_cells = set()  # 병합으로 숨겨진 셀
        for merged in ws.merged_cells.ranges:
            min_r, min_c = merged.min_row, merged.min_col
            max_r, max_c = merged.max_row, merged.max_col
            rowspan = max_r - min_r + 1
            colspan = max_c - min_c + 1
            merge_map[(min_r, min_c)] = (rowspan, colspan)
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    if not (r == min_r and c == min_c):
                        skip_cells.add((r, c))

        # skip행 이후부터 읽기
        all_rows = list(ws.iter_rows(values_only=True))
        data_rows = all_rows[skip:]
        header = [str(v) if v is not None else "" for v in data_rows[0]]
        body = data_rows[1:]

        # 헤더 행 번호 (1-based, openpyxl 기준)
        header_row_idx = skip + 1  # 헤더가 있는 실제 시트 행

        # ※ 포함 행 제거
        filtered_body = []
        filtered_row_indices = []
        for i, row in enumerate(body):
            first = str(row[0]) if row[0] is not None else ""
            if "※" not in first:
                filtered_body.append(row)
                filtered_row_indices.append(header_row_idx + 1 + i)  # 실제 시트 행번호

        # HTML 생성
        table_html = """<div class="qc-table-wrapper notranslate" translate="no">
<table class="qc-table"><thead><tr>"""
        for h in header:
            table_html += f"<th>{h}</th>"
        table_html += "</tr></thead><tbody>"

        for i, (row, sheet_r) in enumerate(zip(filtered_body, filtered_row_indices)):
            table_html += "<tr>"
            for j, val in enumerate(row):
                sheet_c = j + 1  # 1-based col
                if (sheet_r, sheet_c) in skip_cells:
                    continue
                cell_val = str(val).strip() if val is not None else ""
                cell_val = cell_val.replace("None", "").replace("nan", "")
                cell_val = cell_val.replace("(", "<br>(")
                rs, cs = 1, 1
                if (sheet_r, sheet_c) in merge_map:
                    rs, cs = merge_map[(sheet_r, sheet_c)]
                    # rs 조정: ※행이 제거되었으므로 실제 표시되는 행 수만큼만
                    actual_rs = sum(
                        1 for k in range(sheet_r, sheet_r + rs)
                        if k == header_row_idx + 1 or k in filtered_row_indices
                    )
                    rs = max(actual_rs, 1)
                attrs = ""
                if rs > 1:
                    attrs += f" rowspan="{rs}""
                if cs > 1:
                    attrs += f" colspan="{cs}""
                table_html += f"<td{attrs}>{cell_val}</td>"
            table_html += "</tr>"

        table_html += "</tbody></table></div>"
        return table_html
    except Exception as e:
        st.error(f"standard.xlsx 병합 로드 오류: {e}")
        return None


def save_customer_data(df):
    file_path = os.path.join(BASE_DIR, EXCEL_FILE)
    df.to_excel(file_path, index=False)
    load_data.clear()


LOGO_FILENAME = os.path.join(BASE_DIR, "hanjin_logo.png")
logo_base64 = get_image_base64(LOGO_FILENAME)

st.markdown("""
<style>
.header-wrapper {
    display: flex;
    justify-content: space-between;
    align-items: flex-end;
    width: 100%;
    padding: 10px 0;
    border-bottom: 1px solid #f0f2f6;
    margin-bottom: 20px;
}
.logo-container { height: 65px; }
.brand-logo { height: 65px; width: auto; }
.team-name-fixed {
    font-size: 14px;
    font-weight: 600;
    color: rgba(0, 0, 0, 0.5);
    margin-bottom: 5px;
}
.main-title { color: #FF8C00 !important; font-weight: 800; font-size: 1.85rem; }
.customer-title { color: #FF7F50 !important; font-weight: bold; font-size: 1.45rem; margin-top: 30px; margin-bottom: 15px; }
.admin-badge {
    background-color: #FF8C00;
    color: white;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 12px;
    font-weight: bold;
    margin-left: 10px;
}
.qc-table-wrapper {
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
    width: 100%;
}
.qc-table {
    border-collapse: collapse;
    margin-top: 10px;
    font-size: clamp(10px, 2.2vw, 12px);
    border: 1px solid #DEE2E6;
    table-layout: auto;
    white-space: nowrap;
    width: 100%;
}
.qc-table th {
    padding: clamp(4px, 1.5vw, 8px) clamp(6px, 2vw, 12px);
    border: 1px solid #DEE2E6;
    text-align: center !important;
    vertical-align: middle !important;
    background-color: #F8F9FA !important;
    color: #000000 !important;
    font-weight: bold !important;
}
.qc-table td {
    padding: clamp(4px, 1.5vw, 8px) clamp(6px, 2vw, 12px);
    border: 1px solid #DEE2E6;
    text-align: center !important;
    vertical-align: middle !important;
    background-color: white !important;
    color: #000000 !important;
}
.footer-note { font-size: 12.5px; color: #666; margin-top: 15px; font-weight: 500; }
.guide-text { display: none; }
@media (max-width: 768px) {
    .guide-text {
        display: block;
        font-size: 15px;
        font-weight: bold;
        color: #333;
        margin: 15px 0;
        padding: 15px;
        background-color: #fff4e6;
        border-radius: 8px;
        border-left: 5px solid #FF8C00;
        line-height: 1.4;
    }
}
</style>
""", unsafe_allow_html=True)


def render_header():
    if logo_base64:
        logo_img_html = f'<img src="data:image/png;base64,{logo_base64}" class="brand-logo">'
    else:
        logo_img_html = '<div style="color:#ccc; font-size:12px;">[한진철관 로고 미검출]</div>'
    admin_badge = '<span class="admin-badge">🔓 관리자 모드</span>' if st.session_state.is_admin else ""
    st.markdown(f"""
    <div class="header-wrapper">
        <div class="logo-container">{logo_img_html}</div>
        <div class="team-name-fixed">품질기술팀{admin_badge}</div>
    </div>
    """, unsafe_allow_html=True)


def render_admin_login():
    st.sidebar.markdown("---")
    if not st.session_state.is_admin:
        with st.sidebar.expander("🔐 관리자 로그인"):
            pw = st.text_input("비밀번호", type="password", key="admin_pw_input")
            if st.button("로그인", key="admin_login_btn"):
                if pw == ADMIN_PASSWORD:
                    st.session_state.is_admin = True
                    st.rerun()
                else:
                    st.error("비밀번호가 틀렸습니다.")
    else:
        if st.sidebar.button("🔒 관리자 로그아웃"):
            st.session_state.is_admin = False
            st.session_state.edit_idx = None
            st.session_state.show_add_form = False
            st.rerun()


def render_add_form(df):
    st.markdown("### ➕ 고객사 추가")
    cols = df.columns.tolist()
    new_values = {}
    col_pairs = [cols[i:i+2] for i in range(0, len(cols), 2)]
    for pair in col_pairs:
        form_cols = st.columns(2)
        for j, col_name in enumerate(pair):
            new_values[col_name] = form_cols[j].text_input(col_name, key=f"add_{col_name}")
    c1, c2 = st.columns([1, 5])
    if c1.button("저장", key="add_save"):
        if not new_values.get(cols[0], "").strip():
            st.error("고객사명은 필수 입력입니다.")
        else:
            new_row = pd.DataFrame([new_values])
            updated_df = pd.concat([df, new_row], ignore_index=True)
            save_customer_data(updated_df)
            st.session_state.show_add_form = False
            st.success(f"'{new_values[cols[0]]}' 고객사가 추가되었습니다!")
            st.rerun()
    if c2.button("취소", key="add_cancel"):
        st.session_state.show_add_form = False
        st.rerun()


def render_edit_form(df, idx):
    row = df.iloc[idx]
    st.markdown(f"### 수정 중: {row.iloc[0]}")
    cols = df.columns.tolist()
    updated_values = {}
    col_pairs = [cols[i:i+2] for i in range(0, len(cols), 2)]
    for pair in col_pairs:
        form_cols = st.columns(2)
        for j, col_name in enumerate(pair):
            current_val = str(row[col_name]) if pd.notna(row[col_name]) and str(row[col_name]) != "nan" else ""
            updated_values[col_name] = form_cols[j].text_input(col_name, value=current_val, key=f"edit_{col_name}")
    c1, c2 = st.columns([1, 5])
    if c1.button("저장", key="edit_save"):
        for col_name, val in updated_values.items():
            df.at[idx, col_name] = val
        save_customer_data(df)
        st.session_state.edit_idx = None
        st.success("수정이 완료되었습니다!")
        st.rerun()
    if c2.button("취소", key="edit_cancel"):
        st.session_state.edit_idx = None
        st.rerun()


def main():
    render_header()
    st.markdown('<div class="main-title">📋 품질 통합 관리 시스템</div>', unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["📄 고객 사양서", "⚖️ 품질 보증 기준", "🏭 제강사 정보"])

    # ── 탭1: 고객 사양서 ──
    with tab1:
        df_cust = load_data(EXCEL_FILE)
        if df_cust is not None:
            df_cust = df_cust.dropna(subset=[df_cust.columns[0]])
            for col in df_cust.columns:
                df_cust[col] = df_cust[col].astype(str).str.strip()
            customer_list = df_cust.iloc[:, 0].tolist()
            st.sidebar.header("🏢 고객사 목록")
            if st.session_state.is_admin:
                if st.sidebar.button("➕ 고객사 추가", key="open_add_form"):
                    st.session_state.show_add_form = True
                    st.session_state.edit_idx = None
            sel_idx = st.sidebar.radio("업체를 선택하세요:", options=list(range(len(df_cust))), format_func=lambda i: customer_list[i], index=None, key="customer_radio")
            if sel_idx is None and not st.session_state.show_add_form and st.session_state.edit_idx is None:
                st.markdown('<div class="guide-text">좌상단 >> 화살표를 눌러 고객사를 선택 하십시오.</div>', unsafe_allow_html=True)
            if st.session_state.is_admin and st.session_state.show_add_form:
                render_add_form(df_cust)
            elif st.session_state.is_admin and st.session_state.edit_idx is not None:
                render_edit_form(df_cust, st.session_state.edit_idx)
            elif sel_idx is not None:
                row = df_cust.iloc[sel_idx]
                st.markdown(f'<div class="customer-title">■ {row.iloc[0]}</div>', unsafe_allow_html=True)
                if st.session_state.is_admin:
                    a1, a2, _ = st.columns([1, 1, 8])
                    if a1.button("수정", key="edit_btn"):
                        st.session_state.edit_idx = sel_idx
                        st.session_state.show_add_form = False
                        st.rerun()
                    if a2.button("삭제", key="delete_btn"):
                        st.session_state[f"confirm_delete_{sel_idx}"] = True
                    if st.session_state.get(f"confirm_delete_{sel_idx}", False):
                        st.warning(f"**'{row.iloc[0]}'** 고객사를 정말 삭제하시겠습니까?")
                        d1, d2 = st.columns([1, 5])
                        if d1.button("확인 삭제", key="confirm_del"):
                            updated_df = df_cust.drop(index=sel_idx).reset_index(drop=True)
                            save_customer_data(updated_df)
                            st.session_state[f"confirm_delete_{sel_idx}"] = False
                            st.success("삭제되었습니다.")
                            st.rerun()
                        if d2.button("취소", key="cancel_del"):
                            st.session_state[f"confirm_delete_{sel_idx}"] = False
                            st.rerun()
                for i in range(1, len(row.index)):
                    col_n = row.index[i]
                    raw = row.iloc[i]
                    val = str(raw).strip() if pd.notna(raw) and str(raw).strip() not in ("", "nan") else "-"
                    is_sp = any(k in str(col_n) for k in ["특이사항", "주의", "마킹", "포장"])
                    c = "#E63946" if is_sp else "#495057"
                    st.markdown(f"""
                    <div class="notranslate" translate="no" style="display: flex; border: 1px solid #DEE2E6; margin-bottom: -1px;">
                        <div style="background-color: #F8F9FA; width: 85px; min-width: 85px; padding: 10px 4px; font-weight: bold; color: {c}; border-right: 1px solid #DEE2E6; display: flex; align-items: center; justify-content: center; text-align: center; font-size: 12px; line-height: 1.2; word-break: keep-all;">{col_n}</div>
                        <div class="notranslate" translate="no" style="flex: 1; padding: 10px; background-color: white; font-size: 13.5px; line-height: 1.4; color: #212529; font-weight: 500; word-break: break-all;">{val}</div>
                    </div>
                    """, unsafe_allow_html=True)
        render_admin_login()

    # ── 탭2: 품질 보증 기준 (openpyxl 병합셀 정확 재현) ──
    with tab2:
        st.markdown('<div class="customer-title">⚖️ 품질 보증 표준 가이드</div>', unsafe_allow_html=True)
        table_html = load_standard_with_merge("standard.xlsx", skip=5)
        if table_html:
            st.markdown(table_html, unsafe_allow_html=True)
            st.markdown('<div class="footer-note">※ 기타 수요가 요청사항은 별도 협의에 따른다.</div>', unsafe_allow_html=True)

    # ── 탭3: 제강사 정보 ──
    with tab3:
        st.markdown('<div class="customer-title">🏭 제강사 원산지 분류표</div>', unsafe_allow_html=True)
        mill_data = [
            {"코드": "PSC", "제강사": "포스코", "원산지": "대한민국"},
            {"코드": "HDS", "제강사": "현대제철", "원산지": "대한민국"},
            {"코드": "DBS", "제강사": "동부제철", "원산지": "대한민국"},
            {"코드": "DKS", "제강사": "동국씨엠", "원산지": "대한민국"},
            {"코드": "SEAH", "제강사": "세아씨엠", "원산지": "대한민국"},
            {"코드": "TKS", "제강사": "도쿄", "원산지": "일본"},
            {"코드": "NSC", "제강사": "닛테츠", "원산지": "일본"},
            {"코드": "FMS", "제강사": "포모사", "원산지": "베트남"},
            {"코드": "HOA", "제강사": "호아팟", "원산지": "베트남"},
            {"코드": "CHS", "제강사": "중홍", "원산지": "대만"},
            {"코드": "ANF", "제강사": "안펑", "원산지": "중국"},
            {"코드": "BAO", "제강사": "포두", "원산지": "중국"},
            {"코드": "JYE", "제강사": "징예", "원산지": "중국"},
            {"코드": "RSC", "제강사": "일조강철", "원산지": "중국"},
            {"코드": "AGS", "제강사": "안강", "원산지": "중국"},
            {"코드": "DGH", "제강사": "동화", "원산지": "중국"},
            {"코드": "DSH", "제강사": "딩셩", "원산지": "중국"},
            {"코드": "GUF", "제강사": "국풍", "원산지": "중국"},
            {"코드": "HAN", "제강사": "한단", "원산지": "중국"},
            {"코드": "JER", "제강사": "지룬", "원산지": "중국"},
            {"코드": "MSH", "제강사": "보산", "원산지": "중국"},
            {"코드": "SDG", "제강사": "산동", "원산지": "중국"},
            {"코드": "SDS", "제강사": "승덕", "원산지": "중국"},
            {"코드": "SGS", "제강사": "수도", "원산지": "중국"},
            {"코드": "ZHJ", "제강사": "조건", "원산지": "중국"},
            {"코드": "KGM", "제강사": "카이징", "원산지": "중국"},
            {"코드": "LYN", "제강사": "롄강", "원산지": "중국"},
            {"코드": "NTS", "제강사": "신청강", "원산지": "중국"},
            {"코드": "TNT", "제강사": "천철", "원산지": "중국"},
            {"코드": "TSS", "제강사": "당산강철", "원산지": "중국"},
            {"코드": "YAN", "제강사": "연산강철", "원산지": "중국"},
        ]
        df_mill = pd.DataFrame(mill_data)
        search_q = st.text_input("🔍 제강사 명칭 또는 코드 검색", placeholder="예: PSC, 포스코, 중국...", key="mill_search")
        if search_q:
            df_mill = df_mill[
                df_mill["코드"].str.contains(search_q, case=False, na=False) |
                df_mill["제강사"].str.contains(search_q, case=False, na=False) |
                df_mill["원산지"].str.contains(search_q, case=False, na=False)
            ]
        mill_html = '<div class="qc-table-wrapper notranslate" translate="no"><table class="qc-table"><thead><tr><th>코드</th><th>제강사</th><th>원산지</th></tr></thead><tbody>'
        for _, r in df_mill.iterrows():
            o_style = 'style="color:#007BFF; font-weight:bold;"' if r["원산지"] == "대한민국" else ""
            mill_html += f'<tr><td style="font-weight:bold;">{r["코드"]}</td><td>{r["제강사"]}</td><td {o_style}>{r["원산지"]}</td></tr>'
        mill_html += '</tbody></table></div>'
        st.markdown(mill_html, unsafe_allow_html=True)
        st.markdown('<div class="footer-note">※ 제강사 정보는 검색으로 손쉬운 확인이 가능합니다.</div>', unsafe_allow_html=True)


if __name__ == "__main__":
    main()
